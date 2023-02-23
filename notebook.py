import tkinter as tk
from tkinter import ttk
import customtkinter
import sys
import os
from tkinter import *
from PIL import ImageTk, Image
from tkinter import filedialog
import time 
from tkinter.filedialog import askopenfile
import tkinter.messagebox
from  dbc_parser_phase1 import *
import dbc_parser_phase2 
import Arxml_Generator
from tkinter import messagebox
import openpyxl
from tkinter import PhotoImage
import pandas as pd
import time
import CreateDcmarxml
import subprocess
import rte_generator
import c_code_generator
import Merge_code
import CAPLgenerator
import ReportGenerator
import TestCaseExcelGenerator



global excel_file_path
global nextbuttonclicked
nextbuttonclicked=0
global Selectflag

global cwd
cwd= os.getcwd()
print(cwd)
cwd= str(cwd).replace("/","\\")

##############################################################
def close_flash_screen():
    flash.destroy()

# def update_progress_bar():
#     global progress
#     canvas.create_rectangle(0,0, progress, 100, fill="#2b719e",)
#     flash.update()
#     progress += 4
#     if progress < 1000:
#         flash.after(8,update_progress_bar)
#     else:
#         close_flash_screen()

def update_progress_bar():
    global progress
    progress = 0
    progress_bar['value'] = progress
    # execute the batch script using subprocess
    cmd = f'{cwd}\\setup_env_for_raft.bat'
    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
    
    # read the output and update the progress bar
    while True:
        output = process.stdout.readline()
        if output == b'' and process.poll() is not None:
            # terminate the subprocess when the script is done
            process.terminate()
            # close_flash_screen()
            break
        if output:
            # update the progress bar based on the output
            # for example, you can count the number of lines
            # or search for a specific string
            progress += 3
            progress_bar['value'] = progress
            flash.update()
            # close the Tkinter window when progress reaches 100%
            if progress >= progress_bar.cget('maximum'):
                flash.destroy()
                break
    # update the progress bar to 100% when the script is done
    progress_bar['value'] = progress_bar.cget('maximum')


flash = tk.Tk()
flash.geometry("720x382")
flash.title("RAFT")
flash.overrideredirect(True)
screen_width = flash.winfo_screenwidth()
screen_height = flash.winfo_screenheight()
x = (screen_width/2) - (720/2)
y = (screen_height/2) - (382/2)
flash.geometry("+%d+%d" % (x, y))

welcomeLabel = tk.Label (flash, text="RealThingks Automated Framework for Technologies", foreground='black', bg="white", font=("Calibri Light" ,20)) # Engravers MT
welcomeLabel.pack(side="top", fill="both", expand=True)
label = tk.Label(flash, text="Loading Scripts ...", foreground='black', bg="white")
label.pack(side="top", fill="both", expand=True)



# Create a new style for the progress bar
style = ttk.Style()
style.theme_use('default')
style.configure("custom.Horizontal.TProgressbar", background='#2b719e',barcolor='#3b8ed0')

# Create the progress bar with the new style)
progress_bar = ttk.Progressbar(flash,style = "custom.Horizontal.TProgressbar",orient='horizontal', length=200, mode='determinate')
progress_bar.place(height=1)
progress_bar.pack(side="bottom",fill=tk.BOTH)

# load the original image
original_image = PhotoImage(file=f"{cwd}\\media\\background.png")

# resize the image
resized_image = original_image.subsample(5, 5)
# create a label widget to display the resized image
background_label = tk.Label(flash, image=resized_image)
background_label.pack(fill="both", expand=True)

progress = 0
# flash.after(100, update_progress_bar)
flash.after(100, update_progress_bar)

flash.mainloop()
##############################################################

root = Tk()
root.geometry("960x510")
root.title("RAFT")
root.configure(bg='white')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width/2) - (960/2)
y = (screen_height/2) - (510/2)
root.geometry("+%d+%d" % (x, y))

def run():
    root.destroy()

def CheckFileType():
    if ".dbc" in open_file.file_path:
        Screen2.button_1.configure(fg_color= "green")
        Screen2.button_2.configure(fg_color= "transparent")
    elif ".arxml" in open_file.file_path:
        Screen2.button_1.configure(fg_color="transparent")
        Screen2.button_2.configure(fg_color= "green")
    else:
        tkinter.messagebox.showinfo("Warning","You have enter wrong file")

def CheckFileTypeforarxml(filepath):
    if "Dcm_swc.arxml" in filepath:
        pass
    else:
        tkinter.messagebox.showinfo("Warning","You have enter wrong file")

def open_file():
    SetCheckNextButton(1)
    open_file.file_path = filedialog.askopenfilename(filetypes=(("DBC files", "*.dbc"),("ARXML files", "*.arxml")))
    Screen2.file_entry.delete(0, tk.END)
    Screen2.file_entry.insert(0, open_file.file_path)

def open_file_dcm():
    SetCheckNextButton(1)
    open_file_dcm.file_path = filedialog.askopenfilename()
    CheckFileTypeforarxml(open_file_dcm.file_path)
    Screen6.file_entry_dcm.delete(0, tk.END)
    Screen6.file_entry_dcm.insert(0, open_file_dcm.file_path)   

def open_file_groovy():
    SetCheckNextButton(1)
    open_file_groovy.file_path = filedialog.askopenfilename()
    Screen6.open_file_groovy.delete(0, tk.END)
    Screen6.open_file_groovy.insert(0, open_file_groovy.file_path)

def open_file_cfile():
    SetCheckNextButton(1)
    open_file_cfile.file_path = filedialog.askopenfilename()
    Screen9.file_entry.delete(0, tk.END)
    Screen9.file_entry.insert(0, open_file_cfile.file_path)

def onClick():
    tkinter.messagebox.showinfo("Info.", " A .dbc file is an important component in the development of CAN-based systems, as it helps developers to define and understand the messages and signals being transmitted on the network.exp. FasSysCan_P1_V2.1_Dbg.dbc")
    
def onClick2():
    tkinter.messagebox.showinfo("Info.", "A .arxml files contain data that describes the software components, interfaces, and dependencies of the AUTOSAR system, and they are used in various phases of the development process, from system design to implementation and testing.exp.DCCB1_Ecu_Extract_2022_07a0_updated.arxml")

def SelectedVariant(event=None):
    SelectedVariant.variant = Screen3.variable.get()    
    print(SelectedVariant.variant)

# Create a button to check the state of the checkbox
def ColorSelectall(sheet):
    if Screen4.var1.get() == 1:
    
        print("Checkbox is selected")   
    else:
        print("Checkbox is not selected")

def getPathofFile(file_name):
    file_path = os.path.abspath(file_name)
    return file_path

def createExcelSheet():
    global isExcelcreated
    isExcelcreated = 1
    dbc_parser_phase2.option_selected(SelectedVariant.variant)
    Rxpath = f"{cwd}\\internal\\excelfiles\\rx_list_{SelectedVariant.variant}.xlsx"
    TxPath = f"{cwd}\\internal\\excelfiles\\tx_list_{SelectedVariant.variant}.xlsx"

    RxExcelSheet = pd.read_excel(Rxpath)
    TxExcelSheet = pd.read_excel(TxPath)
    selectedvariantExcelSheet = pd.concat([TxExcelSheet,RxExcelSheet])  
    selectedvariantExcelSheet = selectedvariantExcelSheet.loc[:, ~selectedvariantExcelSheet.columns.str.contains('^Unnamed')]
    selectedvariantExcelSheet.to_excel(f"{cwd}\\internal\\excelfiles\\all_{SelectedVariant.variant}.xlsx")
    time.sleep(3)

def CodeGeneration():
    rte_generator.CreateRTEFile(open_file_cfile.file_path)
    c_code_generator.CcodeGenerator(SelectedVariant.variant)
    Merge_code.MergecfileandDEfines(open_file_cfile.file_path)


def CaplGenRule(variant):
    filterType ='Rx'
    if filterType =='Tx':
        CAPLgenerator.CaplGenerator(variant,filtetype='Tx')
    elif filterType =='Tx':
        CAPLgenerator.CaplGenerator(variant,filtetype='Rx')
        

# define a function to update the state of the checkboxes
def update_checkboxes():
    if Screen4.var1.get() == 1:
        Screen4.var2.set(0)
        Screen4.var3.set(0)
        Screen4.var4.set(0)
        Screen4.var5.set(0)
    elif Screen4.var2.get() == 1:
        Screen4.var1.set(0)
        Screen4.var2.set(0)
        Screen4.var3.set(0)
        Screen4.var4.set(0)
        Screen4.var5.set(0)
    elif Screen4.var3.get() == 1:
        Screen4.var1.set(0)
        Screen4.var2.set(0)
        Screen4.var3.set(0)
        Screen4.var4.set(0)
        Screen4.var5.set(0)
    elif Screen4.var4.get() == 1:
        Screen4.var1.set(0)
        Screen4.var2.set(0)
        Screen4.var3.set(0)
        Screen4.var5.set(0)
    elif Screen4.var5.get() == 1:
        Screen4.var1.set(0)
        Screen4.var2.set(0)
        Screen4.var3.set(0)
        Screen4.var4.set(0)

def filterSelected(selection):
    if selection =="AllTx":
        Direction = 1
        txsignalpath = f"{cwd}\\internal\\excelfiles\\tx_list_{SelectedVariant.variant}.xlsx"
        Arxml_Generator.arxmlgenerator(txsignalpath,Direction)
    elif selection =="AllRx":
        Direction = 0
        rxsignalpath = f"{cwd}\\internal\\excelfiles\\rx_list_{SelectedVariant.variant}.xlsx"
        Arxml_Generator.arxmlgenerator(rxsignalpath,Direction)
    elif selection == "SelectAll":
        pass
    elif selection == "AllCyclic":
        pass
    elif selection == "AllSpontaneous":
        pass
    else:
        pass
   
def allTxsignals():
    Direction = 1
    txsignalpath = f"{cwd}\\internal\\excelfiles\\tx_list_{SelectedVariant.variant}.xlsx"
    # """
    # Colors the rows in the given treeview widget that contain the specified substring
    # in the third column with the given color.

    # Args:
    #     tree: The treeview widget to color the rows in.
    #     color: The color to use for the background of the rows.
    #     substring: The substring to match against in the third column of the rows.
    # """
    # for child in tree.get_children():
    #     if tree.item(child)['values'][7] and value in tree.item(child)['values'][7]:
    #         tree.tag_configure('colored', background=color)
    #         tree.item(child, tags=('colored',))
    #     else:
    #         tree.tag_configure('colored', background='')
    #         tree.item(child, tags=())
    Arxml_Generator.arxmlgenerator(txsignalpath,Direction)

def ColortheSignal(tree,color,value,original_color,checked,FilterType):
    """
    Colors the rows in the given treeview widget that contain the specified substring
    in the third column with the given color.

    Args:
        tree: The treeview widget to color the rows in.
        color: The color to use for the background of the rows.
        original_color: The original color of the rows.
        substring: The substring to match against in the third column of the rows.
        checked: A boolean indicating whether the checkbox is checked or not.
    """
    if FilterType == "Tx":
        for child in tree.get_children():
            if tree.item(child)["values"][7] and value in tree.item(child)['values'][7]:
                if checked:
                    tree.item(child, tags=("tag1",))
                    tree.tag_configure("tag1", background=color)
                else:
                    tree.item(child, tags=("tag2",))
                    tree.tag_configure("tag2", background=original_color)
            else:
                tree.item(child, tags=("tag3",))
                tree.tag_configure("tag3", background=original_color)
    if FilterType == "Rx":
        for child in tree.get_children():
            if tree.item(child)["values"][8] and value in tree.item(child)['values'][8]:
                if checked:
                    tree.item(child, tags=("tag1",))
                    tree.tag_configure("tag1", background=color)
                else:
                    tree.item(child, tags=("tag2",))
                    tree.tag_configure("tag2", background=original_color)
            else:
                tree.item(child, tags=("tag3",))
                tree.tag_configure("tag3", background=original_color)    
    if FilterType == "Selectall":
        pass
    # for child in tree.get_children():
    #     if tree.item(child)["values"][8] == value:
    #         if checked:
    #             tree.item(child, tags=("tag1",))
    #             tree.tag_configure("tag1", background=color)
    #         else:
    #             tree.item(child, tags=("tag2",))
    #             tree.tag_configure("tag2", background=original_color)
    #     else:
    #         tree.item(child, tags=("tag3",))
    #         tree.tag_configure("tag3", background=original_color)
    # for child in tree.get_children():
    #     if checked and tree.item(child)['values'][8] and value in tree.item(child)['values'][8]:
    #         tree.item(child, tags=('colored',))
    #         tree.tag_configure('colored', background=color)
    #     else:
    #         tree.item(child, tags=())
    #         tree.tag_configure('', background=original_color)


def allRxsignal():
    Direction = 0
    rxsignalpath = f"{cwd}\\internal\\excelfiles\\rx_list_{SelectedVariant.variant}.xlsx"
    # """
    # Colors the rows in the given treeview widget that contain the specified substring
    # in the third column with the given color.

    # Args:
    #     tree: The treeview widget to color the rows in.
    #     color: The color to use for the background of the rows.
    #     original_color: The original color of the rows.
    #     substring: The substring to match against in the third column of the rows.
    #     checked: A boolean indicating whether the checkbox is checked or not.
    # """
    # for child in tree.get_children():
    #     if checked and tree.item(child)['values'][8] and value in tree.item(child)['values'][8]:
    #         tree.item(child, tags=('colored',))
    #         tree.tag_configure('colored', background=color)
    #     else:
    #         tree.item(child, tags=())
    #         tree.tag_configure('', background=original_color)

    # for child in tree.get_children():
    #     if tree.item(child)['values'][8] and value in tree.item(child)['values'][8]:
    #         tree.tag_configure('colored', background=color)
    #         tree.item(child, tags=('colored',))
    #     else:
    #         tree.tag_configure('colored', background='')
    #         tree.item(child, tags=()) 
    Arxml_Generator.arxmlgenerator(rxsignalpath,Direction)    

def arxmlGenerationBasedonFilter(selection):
    if (selection == 1):
        def Selectallsignal():
            allsignalpath = f"{cwd}\\internal\\excelfiles\\all_{SelectedVariant.variant}.xlsx"
            Arxml_Generator.arxmlgenerator(allsignalpath,2)  
    elif (selection==2):
        pass
    elif (selection ==3):
        pass
    elif(selection==4):
        pass

def On_button_click():
    SetCheckNextButton.value=1
    print(SetCheckNextButton.value)
    Screen1.button_1.configure(fg_color="green")

def run_screen_1():
    if SetCheckNextButton.value:
        for widget in root.winfo_children():
            widget.destroy()
        Screen1()
        SetCheckNextButton(0) 

def AddColortoSelectedSignal(tree,sheet):
    # Configure the row colors
    tree.tag_configure('oddrow', background='#e6f7ff')
    for i in range(1, sheet.max_row + 1):
        if i % 2 == 0:
            tree.tag_has('evenrow', f'${i - 1}')
        else:
            tree.tag_has('oddrow', f'${i - 1}')

def LaunchReport():
    os.startfile(run_screen_12.Report_path)

def SetCheckNextButton(value):
       SetCheckNextButton.value = value 

def run_screen_2():
    
    if SetCheckNextButton.value:
        for widget in root.winfo_children():
            widget.destroy()
        SetCheckNextButton(0)
        Screen2()   
    else:
        messagebox.showinfo("Warning","Kindly select any one")
    
def run_screen_3():
    if SetCheckNextButton.value:
        SetCheckNextButton(0)
        for widget in root.winfo_children():
            widget.destroy()
        Screen3() 
    else:
        messagebox.showinfo("Warning","Please enter path")

def run_screen_4():

    # if Selectflag:
        for widget in root.winfo_children():
            widget.destroy()
        Screen4()

def run_screen_5():
    for widget in root.winfo_children():
        widget.destroy()
    Screen5()

def run_screen_6():
    for widget in root.winfo_children():
        widget.destroy()
    Screen6()
    
def run_screen_7():
    for widget in root.winfo_children():
        widget.destroy()
    Screen7()    
    
def run_screen_8():
    for widget in root.winfo_children():
        widget.destroy()
    Screen8() 

def run_screen_9():
    for widget in root.winfo_children():
        widget.destroy()
    Screen9() 
    
def run_screen_10():
    for widget in root.winfo_children():
        widget.destroy()
    Screen10()
    
def run_screen_11():
    for widget in root.winfo_children():
        widget.destroy()
    Screen11()
    
def run_screen_12():
    run_screen_12.Report_path = ReportGenerator.ReportGen(variant=SelectedVariant.variant,filterType="Rx")
    for widget in root.winfo_children():
        widget.destroy()
    Screen12()    

def Screen1():
    Screen1.button_1 = customtkinter.CTkButton(root, text="CAN Testing Tool",command=On_button_click,border_width=-3,font=("Poppins" ,30),hover_color="green")
    Screen1.button_1.place(relx=0.2, rely=0.4)
    # Load the image
    img_path = f"{cwd}\\media\\Image2.png"
    img = Image.open(img_path)
    img_tk = ImageTk.PhotoImage(img)
    # Create the label and display the image
    label = Label(root, image=img_tk, foreground="black", background="white")
    label.image = img_tk  # Save the image as a global variable to prevent it from being garbage-collected
    label.place(relx=0.2, rely=0.1, anchor='center')
    button_2 = customtkinter.CTkButton(root, text="DIAG Testing Tool",border_width=-3, font=("Poppins" ,30),hover_color="green")
    button_2.place(relx=0.2, rely=0.6)
    button_3 = customtkinter.CTkButton(root, text="NVM Testing Tool",border_width=-3, font=("Poppins" ,30),hover_color="green")
    button_3.place(relx=0.58, rely=0.6)
    button_4 = customtkinter.CTkButton(root, text="RTE Testing Tool",border_width=-3, font=("Poppins" ,30),hover_color="green")
    button_4.place(relx=0.58, rely=0.4)
    button_back = customtkinter.CTkButton(root, text="BACK", command=back_to_main_screen,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=run_screen_2,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')
    button_next.place(relx=0.9, rely=0.9, anchor='center')

def Screen2():
    welcomeLabel_1 = Label (root, text="Select input file: (DBC OR ARXML)", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.25, anchor='center')
    Screen2.file_entry = tk.Entry(root, font=("Poppins" ,15),width=90)
    # Screen2.file_entry.grid(row=5,column=1)
    Screen2.file_entry.place(relx=0.5, rely=0.4, anchor="center")
    Screen2.file_entry.config(width=60)
    # Load the image
    img_path = f"{cwd}\\media\\Image2.png"
    img = Image.open(img_path)
    img_tk = ImageTk.PhotoImage(img)
    # Create the label and display the image
    label = Label(root, image=img_tk, foreground="black", background="white")
    label.image = img_tk  # Save the image as a global variable to prevent it from being garbage-collected
    label.place(relx=0.2, rely=0.1, anchor='center')
    button_o = customtkinter.CTkButton(root,text="Browse", command=lambda:(open_file(),CheckFileType()),border_width=-3, font=("Poppins" ,20), height=19, width=100)
    button_o.place(relx=0.8, rely=0.4, anchor="center",)
    Screen2.button_1 = customtkinter.CTkButton(root, text="DBC (Legacy)", command=onClick,border_width=-3, font=("Poppins" ,25))
    Screen2.button_1.place(relx=0.17, rely=0.7)
    Screen2.button_2 = customtkinter.CTkButton(root, text="ARXML (Autosar)", command=onClick2, border_width=-3,font=("Poppins" ,25))
    Screen2.button_2.place(relx=0.7, rely=0.7)
    button_back = customtkinter.CTkButton(root, text="BACK", command=lambda:(SetCheckNextButton(1),run_screen_1()),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=run_screen_3,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')

def Screen3():
    welcomeLabel_1 = Label (root, text="Select  ECU Instance/Variant", bg="white", foreground='black', font=("Poppins", 25))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.25, anchor='center') 
    # Load the image
    img_path = f"{cwd}\\media\\Image2.png"
    img = Image.open(img_path)
    img_tk = ImageTk.PhotoImage(img)
    # Create the label and display the image
    label = Label(root, image=img_tk, foreground="black", background="white")
    label.image = img_tk  # Save the image as a global variable to prevent it from being garbage-collected
    label.place(relx=0.2, rely=0.1, anchor='center')
    options = file_path(open_file.file_path)
    Screen3.variable = tk.StringVar(root)
    Screen3.variable.set("Choose an option")     
    dropdown = tk.OptionMenu(root, Screen3.variable, *options, command=SelectedVariant)
    dropdown.config(height=1,width=15)
    dropdown.pack()
    dropdown.configure(background='#3b8ed0',foreground='white')
    dropdown.place(relx=0.5, rely=0.5, anchor="center")
    font = ('Poppins', 20)
    dropdown['font'] = font
    dropdown["menu"].config(font=font)
    # Added button to navigate on "NEXT" or "BACK" page
    button_back = customtkinter.CTkButton(root, text="BACK", command=lambda: (SetCheckNextButton(1),run_screen_2()), border_width=-3,font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=lambda: (createExcelSheet(),run_screen_4(),dbc_parser_phase2.variant_cyclic_signals(SelectedVariant.variant)),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')

def Screen4():
    
    # Create a new frame
    frame = tk.Frame(root,height=50,width=50)
    frame.pack(fill=tk.BOTH, expand=True)
    excel_file_path = f"{cwd}\\internal\\excelfiles\\all_{SelectedVariant.variant}.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['Sheet1']
    rows = sheet.max_row
    tree = ttk.Treeview(frame)
    columns = []
    # Get the number of active columns in the sheet
    for column in sheet.columns:
        if column[0].value is not None:
            columns.append(column[0].value)
 
    tree['columns'] = columns
  
    # Set the width and heading for each active column
    for col in columns:
        if col == columns[0]:
            tree.column(col, width=200, minwidth=200, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)
        elif col == columns[1]:
            tree.column(col, width=200, minwidth=200, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)
        else:
            max_width = 0
            for row in sheet[columns.index(col) +1]:
                if row.value is not None:
                    max_width = max(max_width, len(str(row.value)))
            tree.column(col, width=max_width*5, minwidth=max_width*3, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)

    tree.pack(padx=100,pady=120,fill=tk.BOTH, expand=True)
    # Insert the values of the active columns into the treeview, skipping the header row
    header_row = True
    # checkSignal = []
    chk = []
    var_dict = {} 
    for i in range(1, sheet.max_row + 1):
        values = []
        for j, column in enumerate(sheet.columns):
            if column[0].value is not None:
                values.append(sheet.cell(row=i, column=j+1).value)

        if header_row:
            # Add the new checkbox column to the header row
            header_row = False
            columns.insert(0, "Select")
            tree.column("#0", width=50, minwidth=50, stretch=tk.NO)
            tree.heading("#0", text="Select", anchor=tk.W)
            # tree.insert("#0",text="1")
            # tree.insert('','#0',text='',values=values)
            continue
        # Create a new BooleanVar for each checkbox
        var = BooleanVar()
        chk = tk.Checkbutton(tree, variable=var,onvalue=1, offvalue=0)
        # Insert the new checkbox as the first column in the row
        tree.insert('', 'end', text="", values=values, tags=[var])

    # Add scrollbar
    h_scrollbar = tk.Scrollbar(tree, orient="horizontal", command=tree.xview)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    tree.configure(xscrollcommand=h_scrollbar.set)
    # Add scrollbar
    v_scrollbar = tk.Scrollbar(tree, orient="vertical", command=tree.yview)
    v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree.configure(yscrollcommand=v_scrollbar.set)
    button_next = customtkinter.CTkButton(frame, text="NEXT", command=run_screen_5,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.pack(side=tk.LEFT,padx= 15,pady=1)
    button_next.place(relx=0.9, rely=0.9, anchor='center')
    # Rearrange next and back buttons below the scrollbar
    button_back = customtkinter.CTkButton(frame, text="BACK",command=lambda: (SetCheckNextButton(1),run_screen_3()),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.pack(side=tk.RIGHT,padx= 15,pady=1)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    # adding buttons
    Screen4.var1 = tk.BooleanVar()
    Screen4.chk1 = tk.Checkbutton(frame, text="Select All", variable=Screen4.var1, onvalue=1,offvalue=0)
    Screen4.chk1.pack(side=tk.TOP,before=tree,)
    Screen4.chk1.configure(command=arxmlGenerationBasedonFilter(1))
    Screen4.chk1.place(relx=0.1, rely=0.1, anchor='center',)

    Screen4.var2 = tk.BooleanVar()
    Screen4.chk2 = tk.Checkbutton(frame, text="All Tx", variable=Screen4.var2)
    Screen4.chk2.pack(side=tk.TOP,before=tree)
    Screen4.chk2.place(relx=0.2, rely=0.1, anchor='center')
    # Screen4.chk2.configure(command=lambda:(allTxsignals,update_checkboxes()))
    Screen4.chk2.configure(command=lambda:(allTxsignals(),ColortheSignal(tree, color ="#C6E2FF", value=SelectedVariant.variant, original_color = "white", checked = Screen4.var2.get(),FilterType= "Tx")))
    
    Screen4.var3 = tk.BooleanVar()
    Screen4.chk3 = tk.Checkbutton(frame, text="All Rx", variable=Screen4.var3)
    Screen4.chk3.pack(side=tk.TOP,before=tree)
    Screen4.chk3.place(relx=0.4, rely=0.1, anchor='center')
    Screen4.chk3.configure(command=lambda:(allRxsignal(),ColortheSignal(tree, color ="#C6E2FF", value=SelectedVariant.variant, original_color = "white", checked = Screen4.var3.get(),FilterType= "Rx")))
    
    Screen4.var4 = tk.BooleanVar()
    Screen4.chk4 = tk.Checkbutton(frame, text="All Cyclic", variable=Screen4.var4)
    Screen4.chk4.pack(side=tk.TOP,before=tree)  
    Screen4.chk4.place(relx=0.6, rely=0.1, anchor='center')
    
    Screen4.var5 = tk.BooleanVar()
    Screen4.chk5 = tk.Checkbutton(frame, text="All Spontaneous", variable=Screen4.var5)
    Screen4.chk5.pack(side=tk.TOP,before=tree)
    Screen4.chk5.place(relx=0.8, rely=0.1, anchor='center')
 
def Screen5():
    frame = tk.Frame(root,height=50,width=50)
    frame.pack(fill=tk.BOTH, expand=True)

    if Screen4.chk3:
        excel_file_path = f"{cwd}\\internal\\excelfiles\\rx_list_{SelectedVariant.variant}.xlsx"
    elif Screen4.chk2:
        excel_file_path = f"{cwd}\\internal\\excelfiles\\tx_list_{SelectedVariant.variant}.xlsx"


    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['Sheet1']
    rows = sheet.max_row
    tree = ttk.Treeview(frame)
 
    columns = []
    # Get the number of active columns in the sheet
    for column in sheet.columns:
        if column[0].value is not None:
            columns.append(column[0].value)
 
    tree['columns'] = columns
  
    # Set the width and heading for each active column
    for col in columns:
        if col == columns[0]:
            tree.column(col, width=200, minwidth=200, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)
        elif col == columns[1]:
            tree.column(col, width=200, minwidth=200, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)
        else:
            max_width = 0
            for row in sheet[columns.index(col) +1]:
                if row.value is not None:
                    max_width = max(max_width, len(str(row.value)))
            tree.column(col, width=max_width*5, minwidth=max_width*3, stretch=tk.NO)
            tree.heading(col, text=col, anchor=tk.W)
            
    tree.pack(padx=100,pady=120,fill=tk.BOTH, expand=True)
    header_row = True
    for i in range(1, sheet.max_row + 1):
        values = []
        for j, column in enumerate(sheet.columns):
            if column[0].value is not None:
                values.append(sheet.cell(row=i, column=j+1).value)
        if header_row:
            header_row = False
            continue
        var = BooleanVar()
        chk = tk.Checkbutton(tree, variable=var,onvalue=1, offvalue=0)
        tree.insert('', 'end',text=f"{sheet.cell(row=i, column=1).value}", values=tuple(values))

        # Add scrollbar
    h_scrollbar = tk.Scrollbar(tree, orient="horizontal", command=tree.xview)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    tree.configure(xscrollcommand=h_scrollbar.set)

    welcomeLabel_1 = Label (frame, text="Selected signals are as follows",  font=("Poppins", 25))#Creates label widget 
    welcomeLabel_1.pack(side=tk.TOP,before=tree)
    welcomeLabel_1.place(relx=0.5, rely=0.1, anchor='center')

    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_4,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=lambda:(TestCaseExcelGenerator.CreateTestCase(SelectedVariant.variant,filtertype="Rx"),CAPLgenerator.CaplGenerator(SelectedVariant.variant,filtetype="Rx"),run_screen_6()),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')

def Screen6():
    welcomeLabel_1 = Label (root, text="DBC PARSER", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    welcomeLabel_2 = Label (root, text="Enter Path to DCM ARXML file",bg="white", foreground='black', font=("Poppins", 25))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.3, anchor='center')
    welcomeLabel_a = Label (root, text="(Example:C:\git\RAFT\INPUT\EcuConfiguration\Config\ServiceComponents)",bg="white", foreground='black', font=("Poppins", 10))#Creates label widget 
    welcomeLabel_a.grid(row=6, column=9)
    welcomeLabel_a.place(relx=0.5, rely=0.35, anchor='center')

    welcomeLabel_3 = Label (root, text="Enter Path to Groovy file (optional)",bg="white", foreground='black', font=("Poppins", 25))#Creates label widget 
    welcomeLabel_3.grid(row=6, column=9)
    welcomeLabel_3.place(relx=0.5, rely=0.6, anchor='center')

    Screen6.file_entry_dcm = tk.Entry(root, font=("Poppins" ,15),width=70)
    Screen6.file_entry_dcm.place(relx=0.1, rely=0.4,  height=50, width=800)
    Screen6.file_entry_groovy = tk.Entry(root, font=("Poppins" ,15),width=70)
    Screen6.file_entry_groovy.place(relx=0.1, rely=0.65,  height=50, width=800)
    
    button_o = customtkinter.CTkButton(root, text="Browse", command=open_file_dcm,border_width=-3, font=("Poppins" ,25),  height=50, width=150)
    button_o.place(relx=0.8, rely=0.4)
    
    button_o1 = customtkinter.CTkButton(root, text="Browse", command=open_file_groovy,border_width=-3, font=("Poppins" ,25),  height=50, width=150)
    button_o1.place(relx=0.8, rely=0.65)
    
    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_5,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    
    button_next= customtkinter.CTkButton(root, text="NEXT", command=lambda:(run_screen_7(),CreateDcmarxml.Createdcmarxml(open_file_dcm.file_path)),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')

def Screen7():
    welcomeLabel = Label (root, text="DBC Parser", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel.grid(row=6, column=9)
    welcomeLabel.place(relx=0.5, rely=0.15, anchor='center')
    
    welcomeLabel_1 = Label (root, text="Required files are successfully generated in the </output> folder as follows :", bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.25, anchor='center')
    
    welcomeLabel_2 = Label (root, text="RAFT SWC ARXML", bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.45, anchor='center')
    
    welcomeLabel_3 = Label (root, text="Test CAPL script", bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_3.grid(row=6, column=9)
    welcomeLabel_3.place(relx=0.5, rely=0.55, anchor='center')
    
    welcomeLabel_4 = Label (root, text="C code", bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_4.grid(row=6, column=9)
    welcomeLabel_4.place(relx=0.5, rely=0.65, anchor='center')
    
    welcomeLabel_5 = Label (root, text="Groovy Script", bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_5.grid(row=6, column=9)
    welcomeLabel_5.place(relx=0.5, rely=0.75, anchor='center')
    
    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_6,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=run_screen_8, border_width=-3,font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')  

def Screen8():
   
    welcomeLabel_1 = Label (root, text="To do by developer", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    
    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_7,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=run_screen_9,border_width=-3, font=("Poppins" ,25), height=40, width=100,state="disable")
    button_next.place(relx=0.9, rely=0.9, anchor='center') 
    def check_all():
      if Screen8.var1.get() and Screen8.var2.get() and Screen8.var3.get():
        button_next.configure(state='normal')
      else:
        button_next.configure(state='disabled')

    Screen8.var1 = customtkinter.BooleanVar()
    Screen8.chk1 = customtkinter.CTkCheckBox(root, text="Import the CAPL in Restbus                                               ", variable=Screen8.var1,command=check_all)
    Screen8.chk1.pack(side=customtkinter.TOP)
    Screen8.chk1.place(relx=0.5, rely=0.4, anchor='center')
    # Screen8.chk3.configure(command=allRxsignal)
    
    Screen8.var2 = customtkinter.BooleanVar()
    Screen8.chk2 = customtkinter.CTkCheckBox(root, text="Run the test (by clicking start button, Pop up by CAPL", variable=Screen8.var2,command=check_all)
    Screen8.chk2.pack(side=customtkinter.TOP)
    Screen8.chk2.place(relx=0.5, rely=0.5, anchor='center')
    
    Screen8.var3 = customtkinter.BooleanVar()
    Screen8.chk3 = customtkinter.CTkCheckBox(root, text="Run the test (by clicking start button, Pop up by CAPL)", variable=Screen8.var3,command=check_all)
    Screen8.chk3.pack(side=customtkinter.TOP)
    Screen8.chk3.place(relx=0.5, rely=0.6, anchor='center')
  
def Screen9():
    welcomeLabel_1 = Label (root, text="DBC Parser",bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    welcomeLabel_2 = Label (root, text="Enter path to the generated RAFT_SWC.c template file:",bg="white", foreground='black', font=("Poppins", 20))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.35, anchor='center')
    
    Screen9.file_entry = tk.Entry(root, font=("Poppins" ,15),width=70)
    Screen9.file_entry.place(relx=0.1, rely=0.5,  height=50, width=800)
    
    button_o = customtkinter.CTkButton(root, text="Browse", command=open_file_cfile,border_width=-3, font=("Poppins" ,35),  height=50, width=150)
    button_o.place(relx=0.8, rely=0.5)

    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_8,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=lambda:(run_screen_10(),CodeGeneration()),border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')     
    
    
def Screen10():
    welcomeLabel_1 = Label (root, text="DBC PARSER",bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    welcomeLabel_2 = Label (root, text="Information for Developer",bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.35, anchor='center')
    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_9,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_next= customtkinter.CTkButton(root, text="NEXT", command=run_screen_11,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button_next.place(relx=0.9, rely=0.9, anchor='center')     

def Screen11():
    welcomeLabel_1 = Label (root, text="DBC PARSER", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    welcomeLabel_2 = Label (root, text="Enter Path for restbus simulation",bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.3, anchor='center')
    button_o = customtkinter.CTkButton(root, text="Browse", command=open_file, border_width=-3,font=("Poppins" ,25),  height=50, width=150)
    button_o.place(relx=0.8, rely=0.4)
    button_g = customtkinter.CTkButton(root, text="Generate Report", command=run_screen_12 ,font=("Poppins" ,25), border_width=-3, height=40, width=100)
    button_g.place(relx=0.85, rely=0.9, anchor='center')
    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_10, border_width=-3,font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')

    Screen11.file_entry = tk.Entry(root, font=("Poppins" ,15),width=70)
    # Screen2.file_entry.grid(row=5,column=1)
    Screen11.file_entry.place(relx=0.1, rely=0.4,  height=50, width=800)
    button_o = customtkinter.CTkButton(root, text="Browse", command=open_file,border_width=-3, font=("Poppins" ,35),  height=50, width=150)
    button_o.place(relx=0.8, rely=0.4)
    
def Screen12():
    welcomeLabel_1 = Label (root, text="DBC PARSER", bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.15, anchor='center')
    welcomeLabel_2 = Label (root, text="Report Generated Successfully!",bg="white", foreground='black', font=("Poppins", 30))#Creates label widget 
    welcomeLabel_2.grid(row=6, column=9)
    welcomeLabel_2.place(relx=0.5, rely=0.3, anchor='center')
    # welcomeLabel_3 = Label (root, text="HTML Report",bg="white", foreground='black', font=("Poppins", 50))#Creates label widget 
    # welcomeLabel_3.grid(row=6, column=9)
    # welcomeLabel_3.place(relx=0.5, rely=0.5, anchor='center')
    button_back = customtkinter.CTkButton(root, text="Click here", command=LaunchReport, border_width=-3,font=("Poppins" ,20),bg_color="white",fg_color="white", height=40, width=100)
    button_back.place(relx=0.5, rely=0.5, anchor='center')


    button_back = customtkinter.CTkButton(root, text="BACK", command=run_screen_11, border_width=-3,font=("Poppins" ,25), height=40, width=100)
    button_back.place(relx=0.1, rely=0.9, anchor='center')
    button_close = customtkinter.CTkButton(root, text="Close RAFT", command=run, border_width=-3,font=("Poppins" ,25), height=40, width=100)
    button_close.place(relx=0.85, rely=0.9, anchor='center')
    
def back_to_main_screen():
    SetCheckNextButton(0)
    for widget in root.winfo_children():
        widget.destroy()
    main_frame()

def back_to_screen_1():
    for widget in root.winfo_children():
        widget.destroy()
    Screen1()
    
def main_frame():
    button1 = customtkinter.CTkButton(root, text="CREATE NEW AUTOMATION", command=lambda:(SetCheckNextButton(1),run_screen_1()), border_width=-3,font=("Poppins" ,35), height=60, width=600)
    button1.place(relx=0.5, rely=0.65, anchor='center') 

    button = customtkinter.CTkButton(root, text="EXIT", command=run,border_width=-3, font=("Poppins" ,25), height=40, width=100)
    button.place(relx=0.1, rely=0.9, anchor='center')

    # Load the image
    img_path = f"{cwd}\\media\\Image2.png"
    img = Image.open(img_path)
    img_tk = ImageTk.PhotoImage(img)

    # Create the label and display the image
    label = Label(root, image=img_tk, foreground="black", background="white")
    label.image = img_tk  # Save the image as a global variable to prevent it from being garbage-collected
    label.place(relx=0.2, rely=0.1, anchor='center')

    welcomeLabel = Label (root, text="RAFT", foreground='black', bg="white", font=("Poppins" ,25))
    welcomeLabel.grid(row=6, column=9)
    welcomeLabel.place(relx=0.9, rely=0.1, anchor='center')
  
    welcomeLabel_1 = Label (root, text="RealThingks Automated Framework of Testing", bg="white", foreground='black', font=("Poppins", 30))
    welcomeLabel_1.grid(row=6, column=9)
    welcomeLabel_1.place(relx=0.5, rely=0.4, anchor='center')

main_frame()
root.mainloop()